#!/usr/bin/env python3
"""Extract outcome-blind major-FX futures volume/OI from CME workbooks."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook


WORKSPACE = Path(__file__).resolve().parents[2]
DEFAULT_ROOT = WORKSPACE / "02. AlphaFactory" / "external" / "cme_daily_volume"
PRODUCTS = {
    "EC": ("EURUSD", "EURO FX FUTURE"),
    "BP": ("GBPUSD", "BRITISH POUND FUTURE"),
    "J1": ("USDJPY", "JAPANESE YEN FUTURE"),
}
DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")
FILENAME_RE = re.compile(r"^daily_volume_(\d{4})(\d{2})(\d{2})\.xlsx$")


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest().upper()


def sheet_trade_date_candidates(raw: Any) -> set[str]:
    match = DATE_RE.search(str(raw))
    if not match:
        raise ValueError(f"trade date not found: {raw!r}")
    first, second, year = map(int, match.groups())
    if year < 100:
        year += 2000
    candidates: set[str] = set()
    for month, day in ((first, second), (second, first)):
        try:
            candidates.add(datetime(year, month, day).date().isoformat())
        except ValueError:
            pass
    if not candidates:
        raise ValueError(f"invalid sheet trade date: {raw!r}")
    return candidates


def trade_date_from_filename(path: Path) -> str:
    match = FILENAME_RE.fullmatch(path.name)
    if not match:
        raise ValueError(f"invalid official filename: {path.name}")
    return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date().isoformat()


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["CME Group Vol and OI by Product"]
    trade_date = trade_date_from_filename(path)
    sheet_candidates = sheet_trade_date_candidates(sheet.cell(row=2, column=1).value)
    if trade_date not in sheet_candidates:
        raise ValueError(f"filename/sheet date mismatch in {path.name}: {sorted(sheet_candidates)}")
    found: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if str(row[0] or "").strip().upper() != "FX":
            continue
        code = str(row[2] or "").strip().upper()
        if code not in PRODUCTS:
            continue
        try:
            indicator_index = next(
                index for index, value in enumerate(row) if str(value or "").strip().upper() in {"F", "O"}
            )
        except StopIteration:
            continue
        if str(row[indicator_index]).strip().upper() != "F":
            continue
        values = row[indicator_index + 1 : indicator_index + 8]
        if len(values) < 7:
            raise ValueError(f"short value vector in {path.name} for {code}")
        symbol, expected_description = PRODUCTS[code]
        description = str(row[3] or "").strip().upper()
        exchange = str(row[1] or "").strip()
        if description != expected_description or exchange != "Chicago Mercantile Exchange (STATS)":
            raise ValueError(f"identity mismatch in {path.name}: {code}:{exchange}:{description}")
        found[code] = {
            "trade_date": trade_date,
            "symbol": symbol,
            "commodity_code": code,
            "total_volume": int(float(values[4] or 0)),
            "open_interest": int(float(values[6] or 0)),
            "source_file": path.name,
        }
    missing = sorted(set(PRODUCTS) - set(found))
    if missing:
        raise ValueError(f"missing products in {path.name}: {missing}")
    return [found[code] for code in PRODUCTS]


def assert_unique_rows(rows: list[dict[str, Any]]) -> None:
    keys = [(row["trade_date"], row["symbol"]) for row in rows]
    if len(keys) != len(set(keys)):
        raise ValueError("duplicate trade-date/symbol rows after filename-authority parsing")


def extract(root: Path) -> dict[str, Any]:
    manifest_path = root / "source_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("selection", {}).get("holdout_2024_2025_acquired") is not False:
        raise ValueError("sealed holdout contract violated")
    rows: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []
    for index, record in enumerate(manifest["files"], start=1):
        path = WORKSPACE / record["path"]
        actual_sha = sha256_file(path)
        if actual_sha != str(record["sha256"]).upper():
            raise ValueError(f"source hash mismatch: {path}")
        try:
            rows.extend(parse_workbook(path))
        except Exception as exc:  # source-density evidence must preserve every failure
            failures.append({"path": record["path"], "error": str(exc)})
        if index % 250 == 0 or index == len(manifest["files"]):
            print(f"CME_DAILY_FX_EXTRACT progress={index}/{len(manifest['files'])}")
    rows.sort(key=lambda row: (row["trade_date"], row["symbol"]))
    assert_unique_rows(rows)
    output_csv = root / "fx_participation.csv"
    temporary_csv = output_csv.with_suffix(".csv.tmp")
    with temporary_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]) if rows else [])
        if rows:
            writer.writeheader()
            writer.writerows(rows)
    os.replace(temporary_csv, output_csv)
    dates = sorted({row["trade_date"] for row in rows})
    by_symbol = {symbol: sum(row["symbol"] == symbol for row in rows) for symbol, _ in PRODUCTS.values()}
    payload = {
        "schema_version": "cme_daily_fx_participation.v1",
        "generated_at_utc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_manifest_path": manifest_path.relative_to(WORKSPACE).as_posix(),
        "source_manifest_sha256": sha256_file(manifest_path),
        "output_path": output_csv.relative_to(WORKSPACE).as_posix(),
        "output_sha256": sha256_file(output_csv),
        "date_first": dates[0] if dates else None,
        "date_last": dates[-1] if dates else None,
        "date_count": len(dates),
        "row_count": len(rows),
        "rows_by_symbol": by_symbol,
        "failures": failures,
        "price_outcomes_accessed": False,
        "research_authorized": False,
    }
    profile_path = root / "fx_participation_profile.json"
    profile_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return payload


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    payload = extract(args.root.resolve())
    print(json.dumps(payload, indent=2))
    return 0 if not payload["failures"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
