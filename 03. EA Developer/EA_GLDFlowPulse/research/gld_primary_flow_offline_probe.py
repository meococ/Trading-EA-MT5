#!/usr/bin/env python3
"""Frozen SPDR primary creation/redemption probe versus matched XAU momentum."""

from __future__ import annotations

import argparse
import bisect
import hashlib
import json
import math
import os
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import MetaTrader5 as mt5
import numpy as np
import pandas as pd
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[2]
HYPOTHESIS_ID = "HYP-GLDFLOW-XAU-M15-002"
EA_NAME = "EA_GLDFlowPulse"
SYMBOL = "XAUUSD"
PREREG = HERE / "HYP-GLDFLOW-XAU-M15-002_FROZEN_PREREG.md"
WORKBOOK = HERE / "data/US_GLD_Archive_EN.xlsx"
WORKBOOK_SHA256 = "8E7F1DA21C7169D1950F865731817E191E897E650454F9FA37AE5AD1CBD08C38"
FROM_DATE = date(2022, 1, 1)
TO_DATE = date(2024, 12, 31)
WARMUP_DATE = date(2021, 1, 1)
HOLDOUT_START = date(2025, 1, 1)
FROM_UTC = datetime(2021, 12, 1, tzinfo=timezone.utc)
TO_UTC = datetime(2024, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
NY = ZoneInfo("America/New_York")
ENTRY_TIME_NY = time(9, 30)
ATR_PERIOD = 14
STOP_ATR_MULT = 1.5
TARGET_R = 1.5
MAX_HOLD_BARS = 16
RISK_PCT = 0.25
COST_POINTS = {"x1": 82.0, "x1_5": 123.0, "x2": 164.0}
ELAPSED_WEEKS = (TO_DATE - FROM_DATE).days / 7.0
REQUIRED_HEADERS = {
    "Date",
    "Ounces of Gold per Share",
    "Total Ounces of Gold in the Trust",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text or text.upper() in {"N/A", "NA", "-", "NULL"}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def parse_excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    parsed = pd.to_datetime(str(value), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def derive_shares(total_ounces: float, ounces_per_share: float) -> int:
    if total_ounces <= 0 or ounces_per_share <= 0:
        raise ValueError("official ounce fields must be positive")
    raw_shares = total_ounces / ounces_per_share
    return int(math.floor(raw_shares / 100_000.0 + 0.5)) * 100_000


@dataclass(frozen=True)
class FlowRow:
    archive_date: date
    derived_shares: int


def load_flow_rows(path: Path) -> tuple[list[FlowRow], dict[str, Any]]:
    """Load only train/warmup payload cells; future rows expose their date only."""

    if sha256_file(path) != WORKBOOK_SHA256:
        raise RuntimeError("official workbook SHA256 mismatch")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = "US GLD Historical Archive"
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"official archive sheet missing: {sheet_name}")
        sheet = workbook[sheet_name]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        missing = REQUIRED_HEADERS - set(headers)
        if missing:
            raise RuntimeError(f"official archive missing headers: {sorted(missing)}")
        index = {header: headers.index(header) for header in REQUIRED_HEADERS}
        date_index = index["Date"]
        ounce_index = index["Total Ounces of Gold in the Trust"]
        per_share_index = index["Ounces of Gold per Share"]
        rows: list[FlowRow] = []
        holdout_payload_cells_accessed = 0
        for cells in sheet.iter_rows(min_row=2):
            row_date = parse_excel_date(cells[date_index].value)
            if row_date is None:
                continue
            if row_date >= HOLDOUT_START:
                # Do not dereference any non-date cell in a sealed holdout row.
                continue
            if row_date < WARMUP_DATE:
                continue
            total_ounces = parse_number(cells[ounce_index].value)
            ounces_per_share = parse_number(cells[per_share_index].value)
            if total_ounces is None or ounces_per_share is None:
                continue
            rows.append(FlowRow(row_date, derive_shares(total_ounces, ounces_per_share)))
        rows.sort(key=lambda item: item.archive_date)
        if len({row.archive_date for row in rows}) != len(rows):
            raise RuntimeError("duplicate official archive dates in train/warmup slice")
        return rows, {
            "sheet": sheet_name,
            "headers_sha256": hashlib.sha256(
                json.dumps(headers, ensure_ascii=False).encode("utf-8")
            ).hexdigest().upper(),
            "eligible_rows_2021_2024": len(rows),
            "holdout_payload_cells_accessed": holdout_payload_cells_accessed,
        }
    finally:
        workbook.close()


@dataclass(frozen=True)
class FlowSignal:
    archive_date: date
    signal_date: date
    delta_shares: int
    direction: int
    entry_idx: int


def build_flow_events(rows: list[FlowRow]) -> list[tuple[date, int, int]]:
    events: list[tuple[date, int, int]] = []
    for previous, current in zip(rows, rows[1:]):
        if current.archive_date < FROM_DATE or current.archive_date > TO_DATE:
            continue
        delta = current.derived_shares - previous.derived_shares
        if delta == 0:
            continue
        events.append((current.archive_date, delta, 1 if delta > 0 else -1))
    return events


def load_m15() -> pd.DataFrame:
    rates = mt5.copy_rates_range(SYMBOL, mt5.TIMEFRAME_M15, FROM_UTC, TO_UTC)
    if rates is None or len(rates) < 100:
        raise RuntimeError(f"insufficient XAUUSD M15 bars: {mt5.last_error()}")
    frame = pd.DataFrame(rates).drop_duplicates(subset=["time"]).sort_values("time")
    frame["time_utc"] = pd.to_datetime(frame["time"], unit="s", utc=True)
    frame = frame.loc[
        (frame["time_utc"] >= pd.Timestamp(FROM_UTC))
        & (frame["time_utc"] <= pd.Timestamp(TO_UTC))
    ].reset_index(drop=True)
    frame["time_ny"] = frame["time_utc"].dt.tz_convert(NY)
    frame["date_ny"] = frame["time_ny"].dt.date
    return frame


def wilder_atr(frame: pd.DataFrame, period: int = ATR_PERIOD) -> np.ndarray:
    high = frame["high"].to_numpy(float)
    low = frame["low"].to_numpy(float)
    close = frame["close"].to_numpy(float)
    previous = np.r_[np.nan, close[:-1]]
    true_range = np.nanmax(
        np.vstack([high - low, np.abs(high - previous), np.abs(low - previous)]),
        axis=0,
    )
    return pd.Series(true_range).ewm(
        alpha=1.0 / period,
        adjust=False,
        min_periods=period,
    ).mean().to_numpy()


def entry_indices(frame: pd.DataFrame) -> dict[date, int]:
    selected = frame.loc[
        (frame["time_ny"].dt.hour == ENTRY_TIME_NY.hour)
        & (frame["time_ny"].dt.minute == ENTRY_TIME_NY.minute)
    ]
    return {row.date_ny: int(idx) for idx, row in selected.iterrows()}


def bind_events(
    events: Iterable[tuple[date, int, int]], entries: dict[date, int]
) -> list[FlowSignal]:
    trading_dates = sorted(day for day in entries if FROM_DATE <= day <= TO_DATE)
    signals: list[FlowSignal] = []
    used_dates: set[date] = set()
    for archive_date, delta, direction in events:
        position = bisect.bisect_right(trading_dates, archive_date)
        if position >= len(trading_dates):
            continue
        signal_date = trading_dates[position]
        if signal_date in used_dates:
            raise RuntimeError(f"multiple archive events map to {signal_date}")
        used_dates.add(signal_date)
        signals.append(
            FlowSignal(archive_date, signal_date, delta, direction, entries[signal_date])
        )
    return signals


@dataclass
class Trade:
    role: str
    archive_date: str
    signal_date: str
    direction: int
    delta_shares: int
    entry_time_utc: str
    exit_time_utc: str
    risk_points: float
    gross_r: float
    cost_r_x1: float
    exit_reason: str


def simulate_trade(
    frame: pd.DataFrame,
    atr: np.ndarray,
    signal: FlowSignal,
    direction: int,
    point: float,
    role: str,
) -> Trade | None:
    idx = signal.entry_idx
    if idx < ATR_PERIOD + 2 or idx >= len(frame) or not math.isfinite(atr[idx - 1]):
        return None
    entry = float(frame.loc[idx, "open"])
    risk_distance = STOP_ATR_MULT * float(atr[idx - 1])
    if risk_distance <= point:
        return None
    stop = entry - direction * risk_distance
    target = entry + direction * TARGET_R * risk_distance
    end = min(len(frame), idx + MAX_HOLD_BARS)
    end = min(end, idx + sum(frame.loc[idx:end - 1, "date_ny"] == signal.signal_date))
    if end <= idx:
        return None
    exit_idx = idx
    gross_r = 0.0
    reason = "TIME"
    for bar_idx in range(idx, end):
        high = float(frame.loc[bar_idx, "high"])
        low = float(frame.loc[bar_idx, "low"])
        stopped = low <= stop if direction == 1 else high >= stop
        targeted = high >= target if direction == 1 else low <= target
        if stopped:  # Frozen pessimistic same-bar tie handling.
            gross_r = -1.0
            exit_idx = bar_idx
            reason = "STOP"
            break
        if targeted:
            gross_r = TARGET_R
            exit_idx = bar_idx
            reason = "TARGET"
            break
        exit_idx = bar_idx
        if bar_idx == end - 1:
            gross_r = direction * (float(frame.loc[bar_idx, "close"]) - entry) / risk_distance
    return Trade(
        role=role,
        archive_date=signal.archive_date.isoformat(),
        signal_date=signal.signal_date.isoformat(),
        direction=direction,
        delta_shares=signal.delta_shares,
        entry_time_utc=frame.loc[idx, "time_utc"].isoformat(),
        exit_time_utc=frame.loc[exit_idx, "time_utc"].isoformat(),
        risk_points=risk_distance / point,
        gross_r=gross_r,
        cost_r_x1=COST_POINTS["x1"] / (risk_distance / point),
        exit_reason=reason,
    )


def prior_24h_momentum(frame: pd.DataFrame, entry_idx: int) -> int:
    entry_time = frame.loc[entry_idx, "time_utc"]
    cutoff = entry_time - pd.Timedelta(hours=24)
    times = frame["time_utc"].array
    old_idx = int(times.searchsorted(cutoff, side="right") - 1)
    if old_idx < 0 or entry_idx < 1:
        return 0
    change = float(frame.loc[entry_idx - 1, "close"]) - float(frame.loc[old_idx, "close"])
    return 1 if change > 0 else -1 if change < 0 else 0


def evaluate(
    frame: pd.DataFrame, atr: np.ndarray, signals: list[FlowSignal], point: float
) -> tuple[list[Trade], list[Trade]]:
    control: list[Trade] = []
    challenger: list[Trade] = []
    for signal in signals:
        external = simulate_trade(frame, atr, signal, signal.direction, point, "challenger")
        if external is not None:
            challenger.append(external)
        control_direction = prior_24h_momentum(frame, signal.entry_idx)
        if control_direction:
            baseline = simulate_trade(frame, atr, signal, control_direction, point, "control")
            if baseline is not None:
                control.append(baseline)
    return control, challenger


def scenario_values(trades: list[Trade], multiplier: float) -> np.ndarray:
    return np.array(
        [trade.gross_r - trade.cost_r_x1 * multiplier for trade in trades], dtype=float
    )


def summarize_values(values: np.ndarray) -> dict[str, Any]:
    positive = values[values > 0]
    negative = values[values < 0]
    pf_infinite = bool(len(positive) and not len(negative))
    pf = float(positive.sum() / abs(negative.sum())) if len(negative) else None
    equity = np.cumsum(values) if len(values) else np.array([], dtype=float)
    peaks = np.maximum.accumulate(np.r_[0.0, equity])
    drawdown = peaks[1:] - equity if len(equity) else np.array([], dtype=float)
    return {
        "profit_factor": pf,
        "profit_factor_infinite": pf_infinite,
        "net_r": float(values.sum()) if len(values) else 0.0,
        "expectancy_r": float(values.mean()) if len(values) else 0.0,
        "max_drawdown_r": float(drawdown.max()) if len(drawdown) else 0.0,
        "max_drawdown_pct_at_0_25_risk": float(drawdown.max() * RISK_PCT)
        if len(drawdown)
        else 0.0,
    }


def metrics(trades: list[Trade]) -> dict[str, Any]:
    scenarios = {
        name: summarize_values(scenario_values(trades, points / COST_POINTS["x1"]))
        for name, points in COST_POINTS.items()
    }
    by_year: dict[str, float] = {}
    x1 = scenario_values(trades, 1.0)
    for trade, value in zip(trades, x1):
        year = trade.signal_date[:4]
        by_year[year] = by_year.get(year, 0.0) + float(value)
    return {
        "trades": len(trades),
        "trades_per_elapsed_week": len(trades) / ELAPSED_WEEKS,
        "long_trades": sum(1 for trade in trades if trade.direction == 1),
        "short_trades": sum(1 for trade in trades if trade.direction == -1),
        "positive_years_x1": sum(1 for value in by_year.values() if value > 0),
        "by_year_net_r_x1": by_year,
        "scenarios": scenarios,
    }


def comparable_pf(summary: dict[str, Any]) -> float:
    if summary["profit_factor_infinite"]:
        return math.inf
    return float(summary["profit_factor"] or 0.0)


def gates(control: dict[str, Any], challenger: dict[str, Any]) -> dict[str, bool]:
    c1 = control["scenarios"]["x1"]
    h1 = challenger["scenarios"]["x1"]
    h15 = challenger["scenarios"]["x1_5"]
    h2 = challenger["scenarios"]["x2"]
    control_pf = comparable_pf(c1)
    challenger_pf = comparable_pf(h1)
    pf_margin = False if math.isinf(control_pf) else challenger_pf >= control_pf + 0.10
    return {
        "cadence_min": challenger["trades_per_elapsed_week"] >= 2.0,
        "cadence_max": challenger["trades_per_elapsed_week"] <= 5.0,
        "pf_x1": challenger_pf >= 1.35,
        "net_r_x1": h1["net_r"] > 0,
        "expectancy_x1": h1["expectancy_r"] >= 0.10,
        "pf_x1_5": comparable_pf(h15) >= 1.25,
        "pf_x2": comparable_pf(h2) >= 1.00,
        "drawdown": h1["max_drawdown_pct_at_0_25_risk"] <= 5.5,
        "positive_years": challenger["positive_years_x1"] >= 2,
        "long_count": challenger["long_trades"] >= 40,
        "short_count": challenger["short_trades"] >= 40,
        "pf_margin_over_control": pf_margin,
        "net_not_below_control": h1["net_r"] >= c1["net_r"],
    }


def ensure_d_path(path: Path, label: str) -> Path:
    resolved = path.resolve()
    if os.name == "nt" and resolved.drive.upper() != "D:":
        raise RuntimeError(f"{label} must be on D:, got {resolved}")
    return resolved


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--terminal", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()
    terminal_path = ensure_d_path(args.terminal, "portable terminal")
    output_path = ensure_d_path(args.out, "probe output")
    for required in (PREREG, WORKBOOK):
        if not required.is_file():
            raise SystemExit(f"frozen input missing: {required}")
    prereg_sha = sha256_file(PREREG)
    workbook_sha = sha256_file(WORKBOOK)
    if workbook_sha != WORKBOOK_SHA256:
        raise SystemExit("frozen workbook hash mismatch")

    flow_rows, archive_audit = load_flow_rows(WORKBOOK)
    events = build_flow_events(flow_rows)
    if not mt5.initialize(path=str(terminal_path), timeout=60_000, portable=True):
        raise SystemExit(f"MT5 initialize failed: {mt5.last_error()}")
    try:
        terminal = mt5.terminal_info()
        if terminal is None:
            raise RuntimeError(f"terminal_info unavailable: {mt5.last_error()}")
        data_path = Path(str(terminal.data_path)).resolve()
        if os.name == "nt" and data_path.drive.upper() != "D:":
            raise RuntimeError(f"MT5 data_path is not D-portable: {data_path}")
        if not mt5.symbol_select(SYMBOL, True):
            raise RuntimeError(f"symbol_select failed: {mt5.last_error()}")
        symbol = mt5.symbol_info(SYMBOL)
        if symbol is None or symbol.point <= 0:
            raise RuntimeError("XAUUSD point geometry unavailable")
        frame = load_m15()
        if any(day >= HOLDOUT_START for day in frame["date_ny"]):
            raise RuntimeError("holdout XAU bar entered frozen train frame")
        atr = wilder_atr(frame)
        entries = entry_indices(frame)
        signals = bind_events(events, entries)
        control_trades, challenger_trades = evaluate(
            frame, atr, signals, float(symbol.point)
        )
        control = metrics(control_trades)
        challenger = metrics(challenger_trades)
        gate_results = gates(control, challenger)
        verdict = "CONTINUE_TO_EA_BUILD_REVIEW" if all(gate_results.values()) else "KILL_AT_OFFLINE_PROBE"
        payload = {
            "schema_version": "gld_primary_flow_probe.v1",
            "hypothesis_id": HYPOTHESIS_ID,
            "ea_name": EA_NAME,
            "generated_at_utc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "verdict": verdict,
            "promotion_eligible": False,
            "window": {"from": FROM_DATE.isoformat(), "to": TO_DATE.isoformat()},
            "elapsed_calendar_weeks": ELAPSED_WEEKS,
            "frozen_inputs": {
                "prereg_path": str(PREREG.relative_to(ROOT)).replace("\\", "/"),
                "prereg_sha256": prereg_sha,
                "workbook_path": str(WORKBOOK.relative_to(ROOT)).replace("\\", "/"),
                "workbook_sha256": workbook_sha,
            },
            "archive_audit": archive_audit,
            "mt5": {
                "terminal_path": str(terminal_path),
                "data_path": str(data_path),
                "commondata_path": str(terminal.commondata_path),
                "portable": True,
                "file_common_used": False,
                "xau_m15_bars": len(frame),
                "point": float(symbol.point),
            },
            "funnel": {
                "derived_flow_rows_2021_2024": len(flow_rows),
                "nonzero_flow_events_2022_2024": len(events),
                "signals_bound_next_us_trading_day": len(signals),
                "control_trades": len(control_trades),
                "challenger_trades": len(challenger_trades),
            },
            "cost_status": "UNVERIFIED_P99_SPREAD_PROXY_NO_COMMISSION_SLIPPAGE",
            "cost_points": COST_POINTS,
            "control": control,
            "challenger": challenger,
            "gates": gate_results,
            "gate_pass_count": sum(gate_results.values()),
            "gate_total": len(gate_results),
            "trades": {
                "control": [asdict(trade) for trade in control_trades],
                "challenger": [asdict(trade) for trade in challenger_trades],
            },
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = output_path.with_suffix(output_path.suffix + ".tmp")
        temporary.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        os.replace(temporary, output_path)
        print(json.dumps({
            "verdict": verdict,
            "gates": f"{payload['gate_pass_count']}/{payload['gate_total']}",
            "signals": len(signals),
            "trades": len(challenger_trades),
            "tpw": challenger["trades_per_elapsed_week"],
            "pf_x1": challenger["scenarios"]["x1"]["profit_factor"],
            "net_r_x1": challenger["scenarios"]["x1"]["net_r"],
            "out": str(output_path),
        }, separators=(",", ":")))
        return 0
    finally:
        mt5.shutdown()


if __name__ == "__main__":
    raise SystemExit(main())
